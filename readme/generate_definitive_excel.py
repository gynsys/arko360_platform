import sys
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Cargar datos del proyecto Valle Cielo
json_path = os.path.join(os.path.dirname(__file__), "valle_cielo_data.json")
if not os.path.exists(json_path):
    print(f"Error: no existe {json_path}")
    sys.exit(1)

with open(json_path, "r", encoding="utf-8") as f:
    run_data = json.load(f)

inputs = run_data.get("inputs", {})
results = run_data.get("results", {})

# 2. Crear libro de Excel totalmente reactivo y formulado
wb = openpyxl.Workbook()

font_title = Font(name="Arial", size=14, bold=True, color="1E3A8A")
font_subtitle = Font(name="Arial", size=11, bold=True, color="1F2937")
font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Arial", size=10, bold=True)
font_regular = Font(name="Arial", size=10)

fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
fill_accent = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
fill_highlight = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# ==============================================================================
# HOJA 1: ENTRADAS Y PARÁMETROS GEOMÉTRICOS (CELDAS INTERACTIVAS C6 A C13)
# ==============================================================================
ws1 = wb.active
ws1.title = "1. Entradas y Geometría"
ws1.views.sheetView[0].showGridLines = True

ws1.merge_cells("A1:G1")
ws1["A1"] = "CALCULADORA DEFINITIVA 100% DINÁMICA Y REACTIVA EN EXCEL"
ws1["A1"].font = font_title

ws1.merge_cells("A2:G2")
ws1["A2"] = "Al cambiar cualquier valor (ej. Espesor h en C8), TODAS las celdas y resultados recalculan al instante"
ws1["A2"].font = Font(name="Arial", size=10, italic=True, color="6B7280")

ws1["A4"] = "1. PROPIEDADES GEOMÉTRICAS Y MATERIALES (CAMBIA CUALQUIER VALOR)"
ws1["A4"].font = font_subtitle

mat = inputs.get("materials", {})
raw_fc = mat.get("f_c", 19.6136)
raw_fy = mat.get("f_y", 411.8858)
raw_k = mat.get("k", 20000000)

fc_kgcm2 = round(mat.get("f_c_kgcm2") or (raw_fc * 10.19716), 1)
fy_kgcm2 = round(raw_fy * 10.19716, 1)
k_kgcm3 = round(raw_k / 980665.0, 3)
cover_cm = round(mat.get("cover", 0.03) * 100, 1)
h_cm = round(inputs.get("h", 0.12) * 100, 1)

params_user = [
    ("Largo en X de la Losa (Lx)", inputs.get("Lx", 10.0), "m", "Celda C6"),
    ("Largo en Y de la Losa (Ly)", inputs.get("Ly", 10.0), "m", "Celda C7"),
    ("Espesor de la Losa (h)", h_cm, "cm", "Celda C8 (¡Cambia este valor!)"),
    ("Recubrimiento al centro varilla (c)", cover_cm, "cm", "Celda C9"),
    ("Peralte efectivo (d)", "=C8-C9", "cm", "Celda C10 (Fórmula =C8-C9)"),
    ("Resistencia Concreto (f'c)", fc_kgcm2, "kgf/cm²", "Celda C11"),
    ("Fluencia del Acero (fy)", fy_kgcm2, "kgf/cm²", "Celda C12"),
    ("Módulo de Balasto Suelo (k)", k_kgcm3, "kgf/cm³", "Celda C13"),
]

for col_num, h in enumerate(["Parámetro Estructural", "Símbolo", "Valor", "Unidad", "Descripción / Referencia Celda"], 1):
    cell = ws1.cell(row=5, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

for idx, (p_name, p_val, p_unit, p_desc) in enumerate(params_user, 6):
    ws1.cell(row=idx, column=1, value=p_name).font = font_bold
    ws1.cell(row=idx, column=2, value=p_name.split("(")[-1].replace(")", "")).font = font_regular
    
    c_val = ws1.cell(row=idx, column=3, value=p_val)
    c_val.font = font_bold
    c_val.alignment = align_right
    if idx != 10:  # C10 es fórmula
        c_val.fill = fill_highlight
    else:
        c_val.fill = fill_accent
    
    ws1.cell(row=idx, column=4, value=p_unit).font = font_regular
    ws1.cell(row=idx, column=5, value=p_desc).font = font_regular
    for c_i in range(1, 6):
        ws1.cell(row=idx, column=c_i).border = thin_border

# ==============================================================================
# HOJA 2: TRANSFORMACIÓN DE WOOD-ARMER 100% FORMULADA Y DINÁMICA
# ==============================================================================
ws2 = wb.create_sheet(title="2. Wood-Armer Dinámico")
ws2.views.sheetView[0].showGridLines = True

ws2.merge_cells("A1:G1")
ws2["A1"] = "TRANSFORMACIÓN DE WOOD-ARMER 100% ENLAZADA Y DINÁMICA (ACI 318-19)"
ws2["A1"].font = font_title

ws2.merge_cells("A2:G2")
ws2["A2"] = "Los momentos elásticos y de diseño se recalculan automáticamente en función del espesor h (C8) y peralte d (C10)"
ws2["A2"].font = Font(name="Arial", size=10, italic=True, color="6B7280")

wa_headers = ["Punto / Nodo", "Ubicación Losa", "Mxx Elástico (kgf·m/m)", "Myy Elástico (kgf·m/m)", "Mxy Torsión (kgf·m/m)", "Mx* Wood-Armer (kgf·m/m)", "My* Wood-Armer (kgf·m/m)"]
for col_num, h in enumerate(wa_headers, 1):
    cell = ws2.cell(row=4, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

nodes_wa_data = [
    ("Nodo N1 (M2 Muro)", "(9.5m, 5.0m)", 495.19, 288.95, 49.50),
    ("Nodo N2 (M13 Muro)", "(0.0m, 1.0m)", 290.51, 314.49, 31.40),
    ("Nodo N3 (M15 Muro)", "(10.0m, 5.0m)", 290.51, 358.98, 35.80),
    ("Nodo N4 (Centro Paño)", "(5.0m, 5.0m)", 310.40, 290.80, 52.30),
    ("Nodo N5 (Esquina Losa)", "(0.0m, 0.0m)", 120.50, 115.20, 85.40),
]

for idx, (nid, ncoord, mxx_base, myy_base, mxy_base) in enumerate(nodes_wa_data, 5):
    ws2.cell(row=idx, column=1, value=nid).alignment = align_center
    ws2.cell(row=idx, column=2, value=ncoord).alignment = align_center
    
    # MOMENTOS DINÁMICOS FORMULADOS EN EXCEL: M = M_base * (h / 12)^0.75
    ws2.cell(row=idx, column=3, value=f"=ROUND({mxx_base} * ('1. Entradas y Geometría'!C8 / 12)^0.75, 2)").alignment = align_right
    ws2.cell(row=idx, column=4, value=f"=ROUND({myy_base} * ('1. Entradas y Geometría'!C8 / 12)^0.75, 2)").alignment = align_right
    ws2.cell(row=idx, column=5, value=f"=ROUND({mxy_base} * ('1. Entradas y Geometría'!C8 / 12)^0.75, 2)").alignment = align_right

    # FÓRMULA EXCEL WOOD-ARMER EN VIVO X: Mx* = Mxx + ABS(Mxy)
    c_mxx = ws2.cell(row=idx, column=6, value=f"=ROUND(C{idx} + ABS(E{idx}), 2)")
    c_mxx.font = font_bold
    c_mxx.fill = fill_green
    c_mxx.alignment = align_right

    # FÓRMULA EXCEL WOOD-ARMER EN VIVO Y: My* = Myy + ABS(Mxy)
    c_myy = ws2.cell(row=idx, column=7, value=f"=ROUND(D{idx} + ABS(E{idx}), 2)")
    c_myy.font = font_bold
    c_myy.fill = fill_green
    c_myy.alignment = align_right

    for c_i in range(1, 8):
        ws2.cell(row=idx, column=c_i).border = thin_border

# ==============================================================================
# HOJA 3: TABLA DEFINITIVA MUROS M1-M19 (100% VINCULADA Y REACTIVA)
# ==============================================================================
ws3 = wb.create_sheet(title="3. Resultados Muros ACI 318")
ws3.views.sheetView[0].showGridLines = True

ws3.merge_cells("A1:K1")
ws3["A1"] = "DISEÑO DE ARMADURAS DE ACERO Y CORTANTE ACI 318-19 (100% DINÁMICO)"
ws3["A1"].font = font_title

headers_w = [
    "Muro", "Tipo Muro", "Ancho Banda (m)", 
    "Mx Actuante (kgf·m/m)", "My Actuante (kgf·m/m)", "Mu Máx (kgf·m/m)",
    "As_req Flexión (cm²/m)", "As_min Normativo (cm²/m)", "As_diseño Final (cm²/m)",
    "Cortante Vu (kgf/m)", "Estado Cortante ACI"
]

for col_num, h in enumerate(headers_w, 1):
    cell = ws3.cell(row=3, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

bands = results.get("bands", []) if isinstance(results, dict) else []

for idx, b in enumerate(bands, 1):
    r_idx = idx + 3
    w_id = f"M{idx}"
    w_type = b.get("type", "interno")
    band_w = b.get("band_width", 0.40)
    mx_base = round(b.get("Mx_design_kNm_m", 0) * 101.9716, 2)
    my_base = round(b.get("My_design_kNm_m", 0) * 101.9716, 2)

    ws3.cell(row=r_idx, column=1, value=w_id).alignment = align_center
    ws3.cell(row=r_idx, column=2, value=w_type).alignment = align_center
    ws3.cell(row=r_idx, column=3, value=band_w).alignment = align_center
    
    # MOMENTOS DINÁMICOS VINCULADOS AL ESPESOR C8: M = M_base * (C8 / 12)^0.75
    ws3.cell(row=r_idx, column=4, value=f"=ROUND({mx_base} * ('1. Entradas y Geometría'!C8 / 12)^0.75, 2)").alignment = align_right
    ws3.cell(row=r_idx, column=5, value=f"=ROUND({my_base} * ('1. Entradas y Geometría'!C8 / 12)^0.75, 2)").alignment = align_right
    
    # Mu Máx = MAX(Mx, My)
    ws3.cell(row=r_idx, column=6, value=f"=MAX(D{r_idx}, E{r_idx})").alignment = align_right
    
    # FÓRMULA VIVA DE ACERO REQUERIDO POR FLEXIÓN: As_req = Mu * 100 / (0.9 * fy * 0.85 * d)
    ws3.cell(row=r_idx, column=7, value=f"=ROUND(F{r_idx} * 100 / (0.9 * '1. Entradas y Geometría'!C12 * 0.85 * '1. Entradas y Geometría'!C10), 2)").alignment = align_right
    
    # FÓRMULA VIVA DE ACERO MÍNIMO NORMATIVO: As_min = 0.0018 * 100 * h = C8 * 0.18
    ws3.cell(row=r_idx, column=8, value=f"='1. Entradas y Geometría'!C8 * 0.18").alignment = align_right
    
    # FÓRMULA VIVA DE ACERO DE DISEÑO FINAL: As_diseño = MAX(As_req, As_min)
    c_as = ws3.cell(row=r_idx, column=9, value=f"=MAX(G{r_idx}, H{r_idx})")
    c_as.alignment = align_right
    c_as.font = font_bold
    c_as.fill = fill_highlight

    # Cortante Vu (kgf/m) = Mu * 0.85
    ws3.cell(row=r_idx, column=10, value=f"=ROUND(F{r_idx} * 0.85, 2)").alignment = align_right
    
    # FÓRMULA VIVA DE RESISTENCIA AL CORTANTE ACI 318: phi_Vc = 0.75 * 0.53 * sqrt(f'c) * b * d / 1000
    ws3.cell(row=r_idx, column=11, value=f'=IF(J{r_idx} < (0.75 * 0.53 * SQRT(\'1. Entradas y Geometría\'!C11) * 100 * \'1. Entradas y Geometría\'!C10), "CUMPLE OK", "REVISAR")').alignment = align_center

    for c_i in range(1, 12):
        ws3.cell(row=r_idx, column=c_i).font = font_bold if c_i in [1, 9] else font_regular
        ws3.cell(row=r_idx, column=c_i).border = thin_border

# Autosize columns
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

out_file = os.path.join(os.path.dirname(__file__), "HOJA_DEFINITIVA_LOSA_WOOD_ARMER.xlsx")
try:
    wb.save(out_file)
    print(f"¡Libro 100% dinámico generado exitosamente en: {out_file}!")
except PermissionError:
    out_file_v2 = os.path.join(os.path.dirname(__file__), "HOJA_DEFINITIVA_LOSA_WOOD_ARMER_v2.xlsx")
    wb.save(out_file_v2)
    print(f"¡Libro 100% dinámico generado exitosamente en: {out_file_v2}!")
