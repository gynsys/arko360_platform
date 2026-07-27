import sys
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Cargar datos exportados
json_path = os.path.join(os.path.dirname(__file__), "valle_cielo_data.json")
if not os.path.exists(json_path):
    print(f"Error: no existe {json_path}")
    sys.exit(1)

with open(json_path, "r", encoding="utf-8") as f:
    run_data = json.load(f)

inputs = run_data.get("inputs", {})
results = run_data.get("results", {})

# Ejecutar el solver en backend para extraer matrices completas
sys.path.append(os.path.join(os.path.dirname(__file__), "..", "backend"))
from app.schemas.calculadora import SlabModelInput
from app.engine.foundation.facade import FoundationSlabDesigner

input_model = SlabModelInput.model_validate(inputs)
mat = input_model.materials
gr = FoundationSlabDesigner(
    Lx=input_model.geometry.Lx, Ly=input_model.geometry.Ly, h=input_model.geometry.h,
    E=mat.E, nu=mat.nu, k=mat.k, f_c=mat.f_c, f_y=mat.f_y,
    cover=mat.cover, bar_diam=mat.bar_diam,
    gamma_horm=mat.gamma_horm, include_self_weight=True, lambda_aci=1.0,
    band_width_factor=input_model.band_width_factor,
    max_settlement_ratio=input_model.max_settlement_ratio,
    q_adm=mat.q_adm,
    band_width_m=mat.band_width_m,
    custom_mesh_cm2_m=mat.custom_mesh_cm2_m
)
gr.set_mesh(nx=input_model.mesh_nx, ny=input_model.mesh_ny)
for w in input_model.walls:
    openings_dicts = [op.model_dump() if hasattr(op, 'model_dump') else op.dict() for op in w.openings] if w.openings else None
    gr.add_wall(w.x1, w.y1, w.x2, w.y2, w.thickness, w.height, w.density, w.load_factor, w.type, is_plastered=w.is_plastered, openings=openings_dicts)
for c in input_model.columns:
    gr.add_column(c.x, c.y, c.width, c.length, c.height, c.load_kgf, c.id)

gr.solve()
gr.compute_moments()

# Matrices nodales (shape: n_nodes_y, n_nodes_x)
# Convertir momentos de N·m/m a kgf·m/m (factor 0.1019716)
w_grid = gr.w * 1000.0                       # mm
mxx_grid = gr.Mx_raw * 0.1019716              # kgf·m/m
myy_grid = gr.My_raw * 0.1019716              # kgf·m/m
mxy_grid = gr.Mxy * 0.1019716                 # kgf·m/m
mx_wa_grid = gr.Mx * 0.1019716                # kgf·m/m
my_wa_grid = gr.My * 0.1019716                # kgf·m/m

nx, ny = gr.nx, gr.ny
dx, dy = gr.dx, gr.dy
n_nodes_x, n_nodes_y = gr.n_nodes_x, gr.n_nodes_y

# 2. Crear libro de Excel
wb = openpyxl.Workbook()

font_title = Font(name="Arial", size=14, bold=True, color="1E3A8A")
font_subtitle = Font(name="Arial", size=11, bold=True, color="1F2937")
font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Arial", size=10, bold=True)
font_regular = Font(name="Arial", size=10)

fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
fill_accent = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
fill_highlight = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

def add_matrix_sheet(wb_ref, sheet_title, main_heading, data_matrix, unit_str):
    ws = wb_ref.create_sheet(title=sheet_title)
    ws.views.sheetView[0].showGridLines = True
    
    ws.merge_cells("A1:AE1")
    ws["A1"] = f"{main_heading} ({n_nodes_x} x {n_nodes_y} NODOS) [{unit_str}]"
    ws["A1"].font = font_title

    ws.cell(row=3, column=1, value="Y / X (m)").font = font_header
    ws.cell(row=3, column=1).fill = fill_header

    for i in range(n_nodes_x):
        x_val = round(i * dx, 2)
        cell = ws.cell(row=3, column=i + 2, value=x_val)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    for j in range(n_nodes_y):
        y_val = round(j * dy, 2)
        r_idx = j + 4
        cell_y = ws.cell(row=r_idx, column=1, value=y_val)
        cell_y.font = font_bold
        cell_y.alignment = align_center
        cell_y.border = thin_border

        for i in range(n_nodes_x):
            val_ij = round(float(data_matrix[j, i]), 2)
            c_val = ws.cell(row=r_idx, column=i + 2, value=val_ij)
            c_val.font = font_regular
            c_val.alignment = align_right
            c_val.border = thin_border
    return ws

# Eliminar pestaña por defecto y agregar matrices completas
wb.remove(wb.active)
add_matrix_sheet(wb, "1. Deflexiones w(x,y)", "MATRIZ DE DEFLEXIONES NODALES w(x,y)", w_grid, "mm")
add_matrix_sheet(wb, "2. Momento Mxx", "MATRIZ DE MOMENTOS ELÁSTICOS Mxx(x,y)", mxx_grid, "kgf·m/m")
add_matrix_sheet(wb, "3. Momento Myy", "MATRIZ DE MOMENTOS ELÁSTICOS Myy(x,y)", myy_grid, "kgf·m/m")
add_matrix_sheet(wb, "4. Torsión Mxy", "MATRIZ DE MOMENTOS TORSORES Mxy(x,y)", mxy_grid, "kgf·m/m")
add_matrix_sheet(wb, "5. Wood-Armer Mx", "MATRIZ DE DISEÑO WOOD-ARMER Mx(x,y)", mx_wa_grid, "kgf·m/m")
add_matrix_sheet(wb, "6. Wood-Armer My", "MATRIZ DE DISEÑO WOOD-ARMER My(x,y)", my_wa_grid, "kgf·m/m")

# ==============================================================================
# HOJA EVALUADORA POR COORDENADAS (X,Y) CON EXTRACCIÓN Y FÓRMULAS VIVAS
# ==============================================================================
ws_eval = wb.create_sheet(title="7. Evaluador por Coordenadas")
ws_eval.views.sheetView[0].showGridLines = True

ws_eval.merge_cells("A1:G1")
ws_eval["A1"] = "EVALUADOR INTERACTIVO DE MOMENTOS POR COORDENADAS (X, Y)"
ws_eval["A1"].font = font_title

ws_eval.merge_cells("A2:G2")
ws_eval["A2"] = "Ingresa las coordenadas X e Y deseadas para extraer y calcular los momentos paso a paso"
ws_eval["A2"].font = Font(name="Arial", size=10, italic=True, color="6B7280")

ws_eval["A4"] = "1. INGRESO DE COORDENADAS DESEADAS"
ws_eval["A4"].font = font_subtitle

ws_eval["A5"] = "Coordenada X (m):"
ws_eval["C5"] = 5.0
ws_eval["A6"] = "Coordenada Y (m):"
ws_eval["C6"] = 5.0

for r in [5, 6]:
    ws_eval.cell(row=r, column=1).font = font_bold
    ws_eval.cell(row=r, column=3).font = font_bold
    ws_eval.cell(row=r, column=3).fill = fill_highlight
    ws_eval.cell(row=r, column=3).alignment = align_right
    ws_eval.cell(row=r, column=3).border = thin_border

ws_eval["A8"] = "2. ÍNDICES DE BÚSQUEDA NODAL EN LA LA GRILLA (dx = 0.25m, dy = 0.25m)"
ws_eval["A8"].font = font_subtitle

ws_eval["A9"] = "Índice Nodal Columna (i):"
ws_eval["C9"] = f"=ROUND(C5 / {dx}, 0)"
ws_eval["A10"] = "Índice Nodal Fila (j):"
ws_eval["C10"] = f"=ROUND(C6 / {dy}, 0)"

for r in [9, 10]:
    ws_eval.cell(row=r, column=1).font = font_bold
    ws_eval.cell(row=r, column=3).font = font_bold
    ws_eval.cell(row=r, column=3).alignment = align_right
    ws_eval.cell(row=r, column=3).border = thin_border

ws_eval["A12"] = "3. EXTRACTION PASO A PASO DESDE LAS MATRICES NODALES Y FÓRMULAS DE DISEÑO"
ws_eval["A12"].font = font_subtitle

headers_eval = ["Componente / Variable", "Fórmula Excel en Vivo", "Valor Extraído / Calculado", "Unidad", "Descripción y Referencia"]
for col_n, h in enumerate(headers_eval, 1):
    cell = ws_eval.cell(row=13, column=col_n, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

eval_rows = [
    ("Deflexión Vertical w(X,Y)", "=INDEX('1. Deflexiones w(x,y)'!B4:AF44, C10+1, C9+1)", "Extracción matriz deflexión", "mm", "Deflexión elástica nodal de la placa"),
    ("Momento Elástico Mxx", "=INDEX('2. Momento Mxx'!B4:AF44, C10+1, C9+1)", "Extracción matriz Mxx", "kgf·m/m", "Momento flector elástico en dirección X"),
    ("Momento Elástico Myy", "=INDEX('3. Momento Myy'!B4:AF44, C10+1, C9+1)", "Extracción matriz Myy", "kgf·m/m", "Momento flector elástico en dirección Y"),
    ("Momento Torsor Mxy", "=INDEX('4. Torsión Mxy'!B4:AF44, C10+1, C9+1)", "Extracción matriz Mxy", "kgf·m/m", "Momento torsor por distorsión tangencial"),
    ("Wood-Armer Mx* (Diseño X)", "=ROUND(C15 + ABS(C17), 2)", "Fórmula Wood-Armer X", "kgf·m/m", "Mx* = Mxx + |Mxy| (Diseño flexional X ACI)"),
    ("Wood-Armer My* (Diseño Y)", "=ROUND(C16 + ABS(C17), 2)", "Fórmula Wood-Armer Y", "kgf·m/m", "My* = Myy + |Mxy| (Diseño flexional Y ACI)"),
]

for row_i, (var_name, formula_ex, desc_val, unit_s, ref_s) in enumerate(eval_rows, 14):
    ws_eval.cell(row=row_i, column=1, value=var_name).font = font_bold
    ws_eval.cell(row=row_i, column=2, value=formula_ex).font = font_regular
    
    # Celda de valor formulada
    c_v = ws_eval.cell(row=row_i, column=3, value=formula_ex)
    c_v.font = font_bold
    c_v.alignment = align_right
    if row_i in [18, 19]:
        c_v.fill = fill_green
    else:
        c_v.fill = fill_accent
        
    ws_eval.cell(row=row_i, column=4, value=unit_s).font = font_regular
    ws_eval.cell(row=row_i, column=5, value=ref_s).font = font_regular

    for c_idx in range(1, 6):
        ws_eval.cell(row=row_i, column=c_idx).border = thin_border

# Autosize columns
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

out_file = os.path.join(os.path.dirname(__file__), "BENCHMARK_LOSA_VALLE_CIELO_v4.xlsx")
try:
    wb.save(out_file)
    print(f"¡Libro de auditoría V4 generado exitosamente en: {out_file}!")
except PermissionError:
    out_file_v4 = os.path.join(os.path.dirname(__file__), "BENCHMARK_LOSA_VALLE_CIELO_v5.xlsx")
    wb.save(out_file_v4)
    print(f"¡Libro de auditoría V5 generado exitosamente en: {out_file_v4}!")
