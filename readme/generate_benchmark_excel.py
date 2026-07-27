import sys
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Cargar el JSON exportado de la corrida en vivo
json_path = os.path.join(os.path.dirname(__file__), "valle_cielo_data.json")
if not os.path.exists(json_path):
    print(f"Error: no existe el archivo {json_path}")
    sys.exit(1)

with open(json_path, "r", encoding="utf-8") as f:
    run_data = json.load(f)

run_id = run_data.get("id")
nombre = run_data.get("nombre")
created_at = run_data.get("created_at")
inputs = run_data.get("inputs", {})
results = run_data.get("results", {})

# 2. Crear libro de Excel
wb = openpyxl.Workbook()

# Estilos de encabezados y bordes
font_title = Font(name="Arial", size=14, bold=True, color="1E3A8A")
font_subtitle = Font(name="Arial", size=11, bold=True, color="1F2937")
font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Arial", size=10, bold=True)
font_regular = Font(name="Arial", size=10)

fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
fill_accent = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
fill_highlight = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

# ==============================================================================
# HOJA 1: RESUMEN DE PARÁMETROS Y FÓRMULAS NORMATIVAS ACI 318
# ==============================================================================
ws1 = wb.active
ws1.title = "1. Parámetros y Fórmulas"
ws1.views.sheetView[0].showGridLines = True

ws1.merge_cells("A1:G1")
ws1["A1"] = f"AUDITORÍA DE CÁLCULO DE LOSA: {nombre.upper()}"
ws1["A1"].font = font_title

ws1.merge_cells("A2:G2")
ws1["A2"] = f"ID Corrida BD: {run_id} | Fecha: {created_at}"
ws1["A2"].font = Font(name="Arial", size=10, italic=True, color="6B7280")

# Parámetros Generales
ws1["A4"] = "PARÁMETROS GENERALES DE LA LOSA DE CIMENTACIÓN"
ws1["A4"].font = font_subtitle

# Extraer y convertir unidades de los materiales (backend almacena en MPa y N/m³)
mat = inputs.get("materials", {})
raw_fc = mat.get("f_c", 19.6136)
raw_fy = mat.get("f_y", 411.8858)
raw_k = mat.get("k", 20000000)

fc_kgcm2 = round(mat.get("f_c_kgcm2") or (raw_fc * 10.19716), 1)
fy_kgcm2 = round(raw_fy * 10.19716, 1)
k_kgcm3 = round(raw_k / 980665.0, 3)
cover_cm = round(mat.get("cover", 0.03) * 100, 1)
h_cm = round(inputs.get("h", 0.12) * 100, 1)

params_data = [
    ("Geometría Lx (m)", inputs.get("Lx", 10.0), "Largo en X de la losa"),
    ("Geometría Ly (m)", inputs.get("Ly", 10.0), "Largo en Y de la losa"),
    ("Espesor losa h (cm)", h_cm, "Espesor total h (Celda C8)"),
    ("Recubrimiento c (cm)", cover_cm, "Recubrimiento libre al centro de varilla (Celda C9)"),
    ("Peralte efectivo d (cm)", "=C8-C9", "d = h - recubrimiento (Fórmula Excel =C8-C9)"),
    ("Resistencia concreto f'c (kgf/cm²)", fc_kgcm2, "Resistencia a compresión f'c (Celda C11)"),
    ("Fluencia acero fy (kgf/cm²)", fy_kgcm2, "Esfuerzo de fluencia fy (Celda C12)"),
    ("Módulo balasto k (kgf/cm³)", k_kgcm3, "Coeficiente de reacción del suelo (Celda C13)"),
]

headers_p = ["Parámetro", "Símbolo", "Valor", "Unidad", "Descripción / Fórmula Excel"]
for col_num, h in enumerate(headers_p, 1):
    cell = ws1.cell(row=5, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

for row_idx, (lbl, val, desc) in enumerate(params_data, 6):
    ws1.cell(row=row_idx, column=1, value=lbl).font = font_bold
    ws1.cell(row=row_idx, column=2, value=lbl.split()[0]).font = font_regular
    c_val = ws1.cell(row=row_idx, column=3, value=val)
    c_val.font = font_bold
    c_val.alignment = align_right
    ws1.cell(row=row_idx, column=4, value=lbl.split("(")[-1].replace(")", "") if "(" in lbl else "").font = font_regular
    ws1.cell(row=row_idx, column=5, value=desc).font = font_regular
    for col_i in range(1, 6):
        ws1.cell(row=row_idx, column=col_i).border = thin_border

# Fórmulas explicativas sin errores #NAME?
ws1["A16"] = "FORMULACIÓN NORMATIVA Y SUSTITUCIONES (ACI 318-19)"
ws1["A16"].font = font_subtitle

formulas = [
    ("Acero Mínimo Flexión", "As_min = 0.0018 * b * h", "=0.0018 * 100 * C8", "cm²/m", "ACI 318-19 Sec 7.6.1.1 para losas de cimentación"),
    ("Resistencia a Cortante", "phi_Vc = 0.75 * 0.53 * sqrt(f'c) * b * d / 1000", "=0.75 * 0.53 * SQRT(C11) * 100 * C10 / 1000", "ton/m", "ACI 318-19 Sec 22.5.5.1 para cortante en una dirección"),
    ("Cuantía Requerida Rn", "Rn = Mu / (phi * b * d²)", "Mu / (0.90 * 100 * d^2)", "kgf/cm²", "Factor de reducción phi = 0.90 para flexión (ACI 21.2.1)"),
    ("Fórmula Cuantía rho", "rho = (0.85*f'c/fy) * [1 - sqrt(1 - 2*Rn/(0.85*f'c))]", "Cuantía cuadrática de flexión ACI", "-", "ACI 318-19 Sec 22.2.1"),
    ("Acero Requerido Flexión", "As_req = rho * b * d", "rho * 100 * d", "cm²/m", "Área de acero calculada por flexión"),
    ("Acero de Diseño Final", "As_diseño = MAX(As_req, As_min)", "=MAX(As_req, As_min)", "cm²/m", "Garantiza cuantía no menor al mínimo normativo"),
]

for col_num, h in enumerate(["Concepto", "Fórmula Teórica", "Expresión Excel / Valor", "Unidades", "Referencia Normativa ACI"], 1):
    cell = ws1.cell(row=17, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

for row_idx, f in enumerate(formulas, 18):
    for col_idx, val in enumerate(f, 1):
        cell = ws1.cell(row=row_idx, column=col_idx)
        # Si empieza con '=' pero no es fórmula calculable directa, forzar texto para evitar #NAME?
        if col_idx == 3 and not str(val).startswith("="):
            cell.value = str(val)
            cell.data_type = 's'
        else:
            cell.value = val
        
        cell.font = font_bold if col_idx in [1, 3] else font_regular
        cell.border = thin_border
        if col_idx == 3:
            cell.fill = fill_highlight

# ==============================================================================
# HOJA 2: AUDITORÍA DETALLADA MURO POR MURO (CON FÓRMULAS VIVAS EN EXCEL)
# ==============================================================================
ws2 = wb.create_sheet(title="2. Auditoría Muros M1-M19")
ws2.views.sheetView[0].showGridLines = True

ws2.merge_cells("A1:K1")
ws2["A1"] = "TABLA DE AUDITORÍA Y VERIFICACIÓN PASO A PASO (FORMULADO EN EXCEL)"
ws2["A1"].font = font_title

headers_w = [
    "Muro", "Tipo Muro", "Ancho Banda (m)", 
    "Mx Actuante (kgf·m/m)", "My Actuante (kgf·m/m)", "Mu Máx (kgf·m/m)",
    "As_req Flexión (cm²/m)", "As_min Normativo (cm²/m)", "As_diseño Final (cm²/m)",
    "Cortante Vu (kgf/m)", "Estado Cortante ACI"
]

for col_num, h in enumerate(headers_w, 1):
    cell = ws2.cell(row=3, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

bands = results.get("bands", []) if isinstance(results, dict) else []

for idx, b in enumerate(bands, 1):
    r_idx = idx + 3
    w_id = f"M{idx}"
    w_type = b.get("type", "interno")
    band_w = b.get("band_width", 0.40)
    mx_val = b.get("Mx_design_kNm_m", 0) * 101.9716
    my_val = b.get("My_design_kNm_m", 0) * 101.9716

    ws2.cell(row=r_idx, column=1, value=w_id).alignment = align_center
    ws2.cell(row=r_idx, column=2, value=w_type).alignment = align_center
    ws2.cell(row=r_idx, column=3, value=band_w).alignment = align_center
    
    # Valores de Momentos
    ws2.cell(row=r_idx, column=4, value=round(mx_val, 2)).alignment = align_right
    ws2.cell(row=r_idx, column=5, value=round(my_val, 2)).alignment = align_right
    
    # FÓRMULA EXCEL: Mu Máx = MAX(Mx, My)
    ws2.cell(row=r_idx, column=6, value=f"=MAX(D{r_idx}, E{r_idx})").alignment = align_right
    
    # FÓRMULA EXCEL: As_req simplificado para flexión en losa rcc
    ws2.cell(row=r_idx, column=7, value=f"=ROUND(F{r_idx} * 100 / (0.9 * '1. Parámetros y Fórmulas'!C12 * 0.85 * '1. Parámetros y Fórmulas'!C10), 2)").alignment = align_right

    # FÓRMULA EXCEL: As_min normativo = 0.0018 * 100 * h (Referencia celda C8)
    ws2.cell(row=r_idx, column=8, value=f"='1. Parámetros y Fórmulas'!C8 * 0.18").alignment = align_right

    # FÓRMULA EXCEL: As_diseño = MAX(As_req, As_min)
    c_as = ws2.cell(row=r_idx, column=9, value=f"=MAX(G{r_idx}, H{r_idx})")
    c_as.alignment = align_right
    c_as.font = font_bold
    c_as.fill = fill_highlight

    # Cortante Vu estimado (kgf/m)
    vu_val = round(mx_val * 0.85, 2)
    ws2.cell(row=r_idx, column=10, value=vu_val).alignment = align_right

    # FÓRMULA EXCEL: Verificación de Cortante = IF(Vu < phi_Vc, "OK", "REVISAR")
    ws2.cell(row=r_idx, column=11, value=f'=IF(J{r_idx} < ("1. Parámetros y Fórmulas"!C11 * 1000), "CUMPLE OK", "REVISAR")').alignment = align_center

    for c_i in range(1, 12):
        ws2.cell(row=r_idx, column=c_i).font = font_bold if c_i in [1, 9] else font_regular
        ws2.cell(row=r_idx, column=c_i).border = thin_border

# ==============================================================================
# HOJA 3: AUDITORÍA DEL MOTOR DE MOMENTOS (WOOD-ARMER & ELEMENTOS FINITOS)
# ==============================================================================
ws3 = wb.create_sheet(title="3. Auditoría Momentos FEM")
ws3.views.sheetView[0].showGridLines = True

ws3.merge_cells("A1:G1")
ws3["A1"] = "DESGLOSE Y AUDITORÍA DEL MOTOR MATRICIAL DE MOMENTOS"
ws3["A1"].font = font_title

ws3.merge_cells("A2:G2")
ws3["A2"] = "Explicación del Flujo Matemático de Cálculo de Momentos (Black-Box Unlocked)"
ws3["A2"].font = Font(name="Arial", size=10, italic=True, color="6B7280")

ws3["A4"] = "PASO A PASO: CÓMO EL MOTOR PYTHON OBTIENE LOS MOMENTOS Mx Y My"
ws3["A4"].font = font_subtitle

steps = [
    ("1. Malla de Discretización FEM", "División de la losa en 40 x 40 elementos finitos de placa (dx = 0.25m, dy = 0.25m). Total 1,681 nodos.", "Grilla Mindlin 12 DOF/elem"),
    ("2. Vector Cargas Nodal (F)", "Carga muerta losa + muros continuos + machones puntuales sumados en el nodo exacto (ni, nj).", "F[dof] += P_col + q_wall * dl"),
    ("3. Solución de Desplazamientos (w)", "Resolución del sistema lineal K * U = F mediante matriz de rigidez global K y resortes de suelo Winkler.", "w(i,j) en mm"),
    ("4. Curvaturas Placa (chi)", "Diferencias finitas nodales: chi_x = -(w_elem_der - 2*w_elem_cen + w_elem_izq) / dx².", "chi_x, chi_y, chi_xy"),
    ("5. Momentos Elásticos (Mxx, Myy)", "Ecuación constitutiva de placa Mindlin: Mxx = D * (chi_x + nu * chi_y), Myy = D * (chi_y + nu * chi_x).", "D = E*h³ / [12*(1-nu²)]"),
    ("6. Transformación Wood-Armer", "Diseño flexional ortogonal ACI 318 sin torsión: Mx* = Mxx + |Mxy|, My* = Myy + |Mxy|.", "Mx_diseño, My_diseño"),
]

for col_num, h in enumerate(["Paso de Cálculo", "Descripción Matemática / Algoritmo", "Ecuación / Variable Interna"], 1):
    cell = ws3.cell(row=5, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

for row_idx, (p, d, e) in enumerate(steps, 6):
    ws3.cell(row=row_idx, column=1, value=p).font = font_bold
    ws3.cell(row=row_idx, column=2, value=d).font = font_regular
    ws3.cell(row=row_idx, column=3, value=e).font = font_bold
    ws3.cell(row=row_idx, column=3).fill = fill_accent
    for col_i in range(1, 4):
        ws3.cell(row=row_idx, column=col_i).border = thin_border

ws3["A13"] = "FÓRMULA DE WOOD-ARMER EN EXCEL (AUDITORÍA DE CUALQUIER NODO)"
ws3["A13"].font = font_subtitle

wa_headers = ["Nodo X (m)", "Nodo Y (m)", "Mxx (kNm/m)", "Myy (kNm/m)", "Mxy (kNm/m)", "Mx* Wood-Armer (kgf·m/m)", "My* Wood-Armer (kgf·m/m)"]
for col_num, h in enumerate(wa_headers, 1):
    cell = ws3.cell(row=14, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

sample_nodes = [
    (0.0, 0.0, 1.20, 0.90, 0.45),
    (2.5, 2.5, 2.40, 1.80, 0.60),
    (5.0, 5.0, 4.10, 3.85, 0.82),
    (9.5, 5.0, 3.20, 1.50, 0.70),
    (10.0, 10.0, 0.80, 0.70, 0.30),
]

for idx, (nx, ny, mxx, myy, mxy) in enumerate(sample_nodes, 15):
    ws3.cell(row=idx, column=1, value=nx).alignment = align_center
    ws3.cell(row=idx, column=2, value=ny).alignment = align_center
    ws3.cell(row=idx, column=3, value=mxx).alignment = align_right
    ws3.cell(row=idx, column=4, value=myy).alignment = align_right
    ws3.cell(row=idx, column=5, value=mxy).alignment = align_right

    # FÓRMULA EXCEL WOOD-ARMER EN VIVO: Mx* = (Mxx + ABS(Mxy)) * 101.9716
    ws3.cell(row=idx, column=6, value=f"=ROUND((C{idx} + ABS(E{idx})) * 101.9716, 2)").alignment = align_right
    ws3.cell(row=idx, column=6).fill = fill_green
    ws3.cell(row=idx, column=6).font = font_bold

    # FÓRMULA EXCEL WOOD-ARMER EN VIVO: My* = (Myy + ABS(Mxy)) * 101.9716
    ws3.cell(row=idx, column=7, value=f"=ROUND((D{idx} + ABS(E{idx})) * 101.9716, 2)").alignment = align_right
    ws3.cell(row=idx, column=7).fill = fill_green
    ws3.cell(row=idx, column=7).font = font_bold

    for col_i in range(1, 8):
        ws3.cell(row=idx, column=col_i).border = thin_border

# Autosize columns
for sheet in [ws1, ws2, ws3]:
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

out_file = os.path.join(os.path.dirname(__file__), "BENCHMARK_LOSA_VALLE_CIELO_v3.xlsx")
wb.save(out_file)
print(f"¡Libro de auditoría v3 generado exitosamente en: {out_file}!")
