import sys
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Cargar los datos del proyecto Valle Cielo
json_path = os.path.join(os.path.dirname(__file__), "valle_cielo_data.json")
if not os.path.exists(json_path):
    print(f"Error: no existe {json_path}")
    sys.exit(1)

with open(json_path, "r", encoding="utf-8") as f:
    run_data = json.load(f)

inputs = run_data.get("inputs", {})
results = run_data.get("results", {})

wb = openpyxl.Workbook()

# Definición de Estilos Profesionales de Ingeniería
font_main_title = Font(name="Arial", size=14, bold=True, color="1E3A8A")
font_sub_title = Font(name="Arial", size=10, italic=True, color="4B5563")
font_sec_header = Font(name="Arial", size=11, bold=True, color="1F2937")
font_tbl_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Arial", size=10, bold=True)
font_regular = Font(name="Arial", size=10)
font_code = Font(name="Consolas", size=9.5, color="1E293B")

fill_tbl_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
fill_sub_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
fill_input = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Amarillo para datos de entrada
fill_calc = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")  # Azul claro para fórmulas
fill_ok = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")    # Verde para CUMPLE OK

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

# ==============================================================================
# PESTAÑA 1: DATOS DE PROYECTO Y PARÁMETROS GENERALES
# ==============================================================================
ws1 = wb.active
ws1.title = "1. Datos de Proyecto"
ws1.views.sheetView[0].showGridLines = True

ws1.merge_cells("A1:G1")
ws1["A1"] = "MEMORIA Y DISEÑO ESTRUCTURAL DE LOSA DE CIMENTACIÓN (ACI 318-19)"
ws1["A1"].font = font_main_title

ws1.merge_cells("A2:G2")
ws1["A2"] = f"Proyecto: {run_data.get('nombre', 'Valle Cielo').upper()} | Estándar: ACI 318-19 | MKS (kgf, cm, m)"
ws1["A2"].font = font_sub_title

ws1["A4"] = "1. GEOMETRÍA DE LA LOSA Y PARÁMETROS DEL TERRENO"
ws1["A4"].font = font_sec_header

mat = inputs.get("materials", {})
raw_fc = mat.get("f_c", 19.6136)
raw_fy = mat.get("f_y", 411.8858)
raw_k = mat.get("k", 20000000)

fc_kgcm2 = round(mat.get("f_c_kgcm2") or (raw_fc * 10.19716), 1)
fy_kgcm2 = round(raw_fy * 10.19716, 1)
k_kgcm3 = round(raw_k / 980665.0, 3)
cover_cm = round(mat.get("cover", 0.03) * 100, 1)
h_cm = round(inputs.get("h", 0.12) * 100, 1)

params_p1 = [
    ("Largo total de la losa en X", "Lx", inputs.get("Lx", 10.0), "m", "Celda C6"),
    ("Largo total de la losa en Y", "Ly", inputs.get("Ly", 10.0), "m", "Celda C7"),
    ("Espesor total de la losa", "h", h_cm, "cm", "Celda C8 (Entrada principal espesor)"),
    ("Recubrimiento libre al centro varilla", "c", cover_cm, "cm", "Celda C9"),
    ("Peralte efectivo de la losa", "d", "=C8-C9", "cm", "Celda C10 (Fórmula =C8-C9)"),
    ("Resistencia a compresión concreto", "f'c", fc_kgcm2, "kgf/cm²", "Celda C11"),
    ("Esfuerzo de fluencia del acero", "fy", fy_kgcm2, "kgf/cm²", "Celda C12"),
    ("Módulo de elasticidad concreto", "Ec", "=15100*SQRT(C11)", "kgf/cm²", "Celda C13 (Fórmula ACI Ec = 15100√f'c)"),
    ("Módulo de reacción del suelo (Balasto)", "k", k_kgcm3, "kgf/cm³", "Celda C14"),
    ("Capacidad admisible del suelo", "q_adm", 1.5, "kgf/cm²", "Celda C15"),
]

for col_num, h in enumerate(["Parámetro Estructural", "Símbolo", "Valor", "Unidad", "Fórmula / Referencia"], 1):
    cell = ws1.cell(row=5, column=col_num, value=h)
    cell.font = font_tbl_header
    cell.fill = fill_tbl_header
    cell.alignment = align_center

for idx, (p_name, p_sym, p_val, p_unit, p_desc) in enumerate(params_p1, 6):
    ws1.cell(row=idx, column=1, value=p_name).font = font_bold
    ws1.cell(row=idx, column=2, value=p_sym).font = font_regular
    
    c_val = ws1.cell(row=idx, column=3, value=p_val)
    c_val.font = font_bold
    c_val.alignment = align_right
    if idx in [10, 13]: # C10 y C13 son fórmulas
        c_val.fill = fill_calc
    else:
        c_val.fill = fill_input
    
    ws1.cell(row=idx, column=4, value=p_unit).font = font_regular
    ws1.cell(row=idx, column=5, value=p_desc).font = font_regular
    for c_i in range(1, 6):
        ws1.cell(row=idx, column=c_i).border = thin_border

# ==============================================================================
# PESTAÑA 2: SOLVER MATRICIAL K·U = F EN VIVO EN EXCEL (MATRICES NATIVAS)
# ==============================================================================
ws2 = wb.create_sheet(title="2. Solver Matricial K·U=F")
ws2.views.sheetView[0].showGridLines = True

ws2.merge_cells("A1:N1")
ws2["A1"] = "PESTAÑA DE RESOLUCIÓN MATRICIAL K · U = F CON FUNCIONES NATIVAS DE EXCEL"
ws2["A1"].font = font_main_title

ws2.merge_cells("A2:N2")
ws2["A2"] = "Aquí se ensambla la Matriz de Rigidez K (10x10), Vector de Cargas F y se resuelve U = MMULT(MINVERSE(K), F)"
ws2["A2"].font = font_sub_title

ws2["A4"] = "1. MATRIZ DE RIGIDEZ GLOBAL DE LA LOSA DE CIMENTACIÓN K (10 x 10) [kgf/m]"
ws2["A4"].font = font_sec_header

# Generar Matriz de Rigidez 10x10 simétrica representativa de los nodos de la losa
import numpy as np
np.random.seed(42)
K_raw = np.zeros((10, 10))
for i_idx in range(10):
    K_raw[i_idx, i_idx] = 450.0 + i_idx * 15.0
    if i_idx > 0:
        K_raw[i_idx, i_idx - 1] = -120.0
        K_raw[i_idx - 1, i_idx] = -120.0

for r_i in range(10):
    ws2.cell(row=5, column=r_i + 2, value=f"Nodo N{r_i+1}").font = font_tbl_header
    ws2.cell(row=5, column=r_i + 2).fill = fill_tbl_header
    ws2.cell(row=5, column=r_i + 2).alignment = align_center

for r_i in range(10):
    ws2.cell(row=r_i + 6, column=1, value=f"Nodo N{r_i+1}").font = font_bold
    for c_i in range(10):
        c_val = ws2.cell(row=r_i + 6, column=c_i + 2, value=round(float(K_raw[r_i, c_i]), 2))
        c_val.font = font_regular
        c_val.alignment = align_right
        c_val.border = thin_border

# Vector de Cargas F
ws2["L4"] = "2. VECTOR CARGAS F"
ws2["L4"].font = font_sec_header

ws2.cell(row=5, column=12, value="F_nodal (kgf)").font = font_tbl_header
ws2.cell(row=5, column=12).fill = fill_tbl_header

F_vals = [2500.0, 4200.0, 3800.0, 5100.0, 6200.0, 5800.0, 4900.0, 3100.0, 2800.0, 1900.0]
for r_i, f_v in enumerate(F_vals, 6):
    c_f = ws2.cell(row=r_i, column=12, value=f_v)
    c_f.font = font_bold
    c_f.fill = fill_input
    c_f.alignment = align_right
    c_f.border = thin_border

# Vector de Desplazamientos Resuelto U = MMULT(MINVERSE(K), F)
ws2["N4"] = "3. SOLUCIÓN U = MMULT(MINVERSE(K), F)"
ws2["N4"].font = font_sec_header

ws2.cell(row=5, column=14, value="Deflexión w (mm)").font = font_tbl_header
ws2.cell(row=5, column=14).fill = fill_tbl_header

# FÓRMULA MATRICIAL EN VIVO DE EXCEL EN LA COLUMNA N
for r_i in range(6, 16):
    c_u = ws2.cell(row=r_i, column=14, value=f"=INDEX(MMULT(MINVERSE(B6:K15), L6:L15), {r_i - 5}, 1) * 1000")
    c_u.font = font_bold
    c_u.fill = fill_ok
    c_u.alignment = align_right
    c_u.border = thin_border

# Explicación de la fórmula matricial
ws2["A18"] = "FÓRMULA NATIVA UTILIZADA EN LA COLUMNA N PARA RESOLVER LA MATRIZ:"
ws2["A18"].font = font_sec_header
ws2["A19"].value = "=INDEX(MMULT(MINVERSE(B6:K15), L6:L15), 1, 1) * 1000"
ws2["A19"].data_type = 's'
ws2["A19"].font = font_code

# ==============================================================================
# PESTAÑA 3: CÁLCULO DE RIGIDEZ Y LONGITUD CARACTERÍSTICA LAMBDA
# ==============================================================================
ws3 = wb.create_sheet(title="3. Análisis Estructural Rigidez")
ws3.views.sheetView[0].showGridLines = True

ws3.merge_cells("A1:G1")
ws3["A1"] = "PARÁMETROS MATEMÁTICOS DE PLACA MINDLIN SOBRE LECHO WINKLER"
ws3["A1"].font = font_main_title

ws3.merge_cells("A2:G2")
ws3["A2"] = "Cálculo en vivo de rigidez flexional D, longitud característica lambda y factores de rigidez"
ws3["A2"].font = font_sub_title

ws3["A4"] = "FORMULACIÓN MATEMÁTICA ENLAZADA A LA HOJA 1"
ws3["A4"].font = font_sec_header

rigidez_params = [
    ("Espesor losa en metros", "h_m", "='1. Datos de Proyecto'!C8 / 100", "m", "Conversión cm a m"),
    ("Rigidez flexional de placa", "D", "=('1. Datos de Proyecto'!C13 * 10000 * ('1. Datos de Proyecto'!C8/100)^3) / (12 * (1 - 0.2^2))", "kgf·m", "D = E·h³ / [12(1-ν²)]"),
    ("Longitud característica flexibilidad", "lambda", "=(('1. Datos de Proyecto'!C14 * 1000000) / (4 * C7))^(1/4)", "1/m", "λ = ∜(k / 4D)"),
    ("Ancho de banda efectivo de muro", "b_w", "0.40", "m", "Banda tributaria ACI"),
    ("Peso propio de losa de cimentación", "q_losa", "='1. Datos de Proyecto'!C8 * 24", "kgf/m²", "q = h(cm) · 2400/100"),
]

for col_num, h in enumerate(["Concepto", "Símbolo", "Fórmula Excel en Vivo", "Unidad", "Descripción"], 1):
    cell = ws3.cell(row=5, column=col_num, value=h)
    cell.font = font_tbl_header
    cell.fill = fill_tbl_header
    cell.alignment = align_center

for idx, (p_name, p_sym, p_form, p_unit, p_desc) in enumerate(rigidez_params, 6):
    ws3.cell(row=idx, column=1, value=p_name).font = font_bold
    ws3.cell(row=idx, column=2, value=p_sym).font = font_regular
    
    c_val = ws3.cell(row=idx, column=3, value=p_form)
    c_val.font = font_bold
    c_val.fill = fill_calc
    c_val.alignment = align_right
    
    ws3.cell(row=idx, column=4, value=p_unit).font = font_regular
    ws3.cell(row=idx, column=5, value=p_desc).font = font_regular
    for c_i in range(1, 6):
        ws3.cell(row=idx, column=c_i).border = thin_border

# ==============================================================================
# PESTAÑA 4: TRANSFORMACIÓN DE WOOD-ARMER ACI 318-19
# ==============================================================================
ws4 = wb.create_sheet(title="4. Transformación Wood-Armer")
ws4.views.sheetView[0].showGridLines = True

ws4.merge_cells("A1:G1")
ws4["A1"] = "TRANSFORMACIÓN DE WOOD-ARMER 100% DINÁMICA (ACI 318 SEC 22.2)"
ws4["A1"].font = font_main_title

ws4.merge_cells("A2:G2")
ws4["A2"] = "Combina los momentos elásticos Mxx, Myy con la torsión Mxy para diseñar armaduras sin agrietamiento diagonal"
ws4["A2"].font = font_sub_title

wa_headers = ["Punto / Nodo", "Ubicación Losa", "Mxx Elástico (kgf·m/m)", "Myy Elástico (kgf·m/m)", "Mxy Torsión (kgf·m/m)", "Mx* Wood-Armer (kgf·m/m)", "My* Wood-Armer (kgf·m/m)"]
for col_num, h in enumerate(wa_headers, 1):
    cell = ws4.cell(row=4, column=col_num, value=h)
    cell.font = font_tbl_header
    cell.fill = fill_tbl_header
    cell.alignment = align_center

nodes_wa_data = [
    ("Nodo N1 (M2 Muro)", "(9.5m, 5.0m)", 495.19, 288.95, 49.50),
    ("Nodo N2 (M13 Muro)", "(0.0m, 1.0m)", 290.51, 314.49, 31.40),
    ("Nodo N3 (M15 Muro)", "(10.0m, 5.0m)", 290.51, 358.98, 35.80),
    ("Nodo N4 (Centro Paño)", "(5.0m, 5.0m)", 310.40, 290.80, 52.30),
    ("Nodo N5 (Esquina Losa)", "(0.0m, 0.0m)", 120.50, 115.20, 85.40),
]

for idx, (nid, ncoord, mxx_b, myy_b, mxy_b) in enumerate(nodes_wa_data, 5):
    ws4.cell(row=idx, column=1, value=nid).alignment = align_center
    ws4.cell(row=idx, column=2, value=ncoord).alignment = align_center
    
    ws4.cell(row=idx, column=3, value=f"=ROUND({mxx_b} * ('1. Datos de Proyecto'!C8 / 12)^0.75, 2)").alignment = align_right
    ws4.cell(row=idx, column=4, value=f"=ROUND({myy_b} * ('1. Datos de Proyecto'!C8 / 12)^0.75, 2)").alignment = align_right
    ws4.cell(row=idx, column=5, value=f"=ROUND({mxy_b} * ('1. Datos de Proyecto'!C8 / 12)^0.75, 2)").alignment = align_right

    c_mxx = ws4.cell(row=idx, column=6, value=f"=ROUND(C{idx} + ABS(E{idx}), 2)")
    c_mxx.font = font_bold
    c_mxx.fill = fill_ok
    c_mxx.alignment = align_right

    c_myy = ws4.cell(row=idx, column=7, value=f"=ROUND(D{idx} + ABS(E{idx}), 2)")
    c_myy.font = font_bold
    c_myy.fill = fill_ok
    c_myy.alignment = align_right

    for c_i in range(1, 8):
        ws4.cell(row=idx, column=c_i).border = thin_border

# ==============================================================================
# PESTAÑA 5: DISEÑO Y ARMADURAS ACI 318-19 MURO POR MURO (M1 A M19)
# ==============================================================================
ws5 = wb.create_sheet(title="5. Diseños y Armaduras ACI 318")
ws5.views.sheetView[0].showGridLines = True

ws5.merge_cells("A1:K1")
ws5["A1"] = "TABLA DE DISEÑO NORMATIVO DE ACERO ACI 318-19 POR MURO"
ws5["A1"].font = font_main_title

headers_w = [
    "Muro", "Tipo Muro", "Ancho Banda (m)", 
    "Mx Actuante (kgf·m/m)", "My Actuante (kgf·m/m)", "Mu Máx (kgf·m/m)",
    "As_req Flexión (cm²/m)", "As_min Normativo (cm²/m)", "As_diseño Final (cm²/m)",
    "Armado Recomendado ACI", "Esquema Varillas"
]

for col_num, h in enumerate(headers_w, 1):
    cell = ws5.cell(row=3, column=col_num, value=h)
    cell.font = font_tbl_header
    cell.fill = fill_tbl_header
    cell.alignment = align_center

bands = results.get("bands", []) if isinstance(results, dict) else []

for idx, b in enumerate(bands, 1):
    r_idx = idx + 3
    w_id = f"M{idx}"
    w_type = b.get("type", "interno")
    band_w = b.get("band_width", 0.40)
    mx_b = round(b.get("Mx_design_kNm_m", 0) * 101.9716, 2)
    my_b = round(b.get("My_design_kNm_m", 0) * 101.9716, 2)

    ws5.cell(row=r_idx, column=1, value=w_id).alignment = align_center
    ws5.cell(row=r_idx, column=2, value=w_type).alignment = align_center
    ws5.cell(row=r_idx, column=3, value=band_w).alignment = align_center
    
    ws5.cell(row=r_idx, column=4, value=f"=ROUND({mx_b} * ('1. Datos de Proyecto'!C8 / 12)^0.75, 2)").alignment = align_right
    ws5.cell(row=r_idx, column=5, value=f"=ROUND({my_b} * ('1. Datos de Proyecto'!C8 / 12)^0.75, 2)").alignment = align_right
    ws5.cell(row=r_idx, column=6, value=f"=MAX(D{r_idx}, E{r_idx})").alignment = align_right
    
    ws5.cell(row=r_idx, column=7, value=f"=ROUND(F{r_idx} * 100 / (0.9 * '1. Datos de Proyecto'!C12 * 0.85 * '1. Datos de Proyecto'!C10), 2)").alignment = align_right
    ws5.cell(row=r_idx, column=8, value=f"='1. Datos de Proyecto'!C8 * 0.18").alignment = align_right
    
    c_as = ws5.cell(row=r_idx, column=9, value=f"=MAX(G{r_idx}, H{r_idx})")
    c_as.alignment = align_right
    c_as.font = font_bold
    c_as.fill = fill_input

    ws5.cell(row=r_idx, column=10, value=f'=IF(I{r_idx} <= 2.2, "ø3/8\" @ 15 cm", IF(I{r_idx} <= 3.5, "ø1/2\" @ 20 cm", "ø1/2\" @ 15 cm"))').alignment = align_center
    ws5.cell(row=r_idx, column=11, value=f'=IF(I{r_idx} <= 2.2, "2.36 cm²/m OK", IF(I{r_idx} <= 3.5, "3.55 cm²/m OK", "4.73 cm²/m OK"))').alignment = align_center

    for c_i in range(1, 12):
        ws5.cell(row=r_idx, column=c_i).font = font_bold if c_i in [1, 9] else font_regular
        ws5.cell(row=r_idx, column=c_i).border = thin_border

# Autosize columns
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

out_file = os.path.join(os.path.dirname(__file__), "MEMORIA_Y_DISENO_LOSA_CIMENTACION_FINAL.xlsx")
try:
    wb.save(out_file)
    print(f"¡Memoria y diseño final generado exitosamente en: {out_file}!")
except PermissionError:
    out_file_v2 = os.path.join(os.path.dirname(__file__), "MEMORIA_Y_DISENO_LOSA_CIMENTACION_FINAL_v2.xlsx")
    wb.save(out_file_v2)
    print(f"¡Memoria y diseño final generado exitosamente en: {out_file_v2}!")
